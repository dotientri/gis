#!/usr/bin/env python
import os
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from parks.models import BaoCaoSuCo
from datetime import date
import openpyxl
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    # Get incidents from today
    queryset = BaoCaoSuCo.objects.select_related(
        'ma_cong_vien', 'ma_danh_muc', 'ma_nguoi_phu_trach', 'ma_nguoi_bao_cao'
    ).filter(is_archived=False, ngay_tao__date=date.today())
    
    print(f"Testing Excel generation with {queryset.count()} incidents...")
    
    # Remove duplicates
    seen = set()
    unique_incidents = []
    for incident in queryset:
        key = (incident.ma_cong_vien.ma_cong_vien, incident.tieu_de)
        if key not in seen:
            seen.add(key)
            unique_incidents.append(incident)
    
    print(f"✅ Unique incidents: {len(unique_incidents)}")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sự cố công viên'
    
    # Styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Headers
    headers = ['STT', 'Công viên', 'Tiêu đề', 'Mô tả', 'Loại sự cố', 'Mức độ ưu tiên',
               'Trạng thái', 'Người báo cáo', 'Người phụ trách', 'Số xác nhận', 'Ngày tạo']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Data
    for row_num, incident in enumerate(unique_incidents, 2):
        print(f"Processing row {row_num}...")
        print(f"  Công viên: {incident.ma_cong_vien.ten_cong_vien}")
        print(f"  Tiêu đề: {incident.tieu_de}")
        
        row_data = [
            row_num - 1,
            incident.ma_cong_vien.ten_cong_vien,
            incident.tieu_de,
            incident.noi_dung_mo_ta,
            incident.ma_danh_muc.ten_danh_muc if incident.ma_danh_muc else 'N/A',
            incident.get_muc_do_uu_tien_display(),
            incident.get_trang_thai_display(),
            incident.ma_nguoi_bao_cao.ten_dang_nhap if incident.ma_nguoi_bao_cao else 'Anonymous',
            incident.ma_nguoi_phu_trach.ten_dang_nhap if incident.ma_nguoi_phu_trach else 'Chưa phân công',
            incident.so_nguoi_xac_nhan,
            incident.ngay_tao.strftime('%d/%m/%Y %H:%M')
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num == 1:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
    
    # Save
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    print(f"✅ Excel file generated successfully ({len(excel_file.getvalue())} bytes)")
    
except Exception as e:
    print(f"❌ Error: {type(e).__name__}: {e}")
    import traceback
    traceback.print_exc()
