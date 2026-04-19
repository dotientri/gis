from rest_framework import viewsets, status, filters, serializers
from rest_framework.decorators import action, api_view
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q, Count, Avg, Sum
from django.utils import timezone 
from datetime import timedelta
from django.conf import settings
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import json

try:
    import geopandas as gpd
    from shapely.geometry import shape, mapping
    from shapely.validation import make_valid
    HAS_GEOPANDAS = True
except ImportError:
    HAS_GEOPANDAS = False

# Import for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from io import BytesIO
from django.http import HttpResponse

from .permissions import (
    IsAuthenticated, IsAdmin, IsParkManager, IsGISEditor,
    IsParkManagerOrGISEditor, IsInspector, IsOwnerOrAdmin,
    CanCreatePark, CanRateAndReview, CanReportIncident,
    IsGuest, IsUser, IsManagerOrAdmin, CanManageAmenities,
    CanHandleIncident, CanCreateEvent, IsReadOnly, CanUpdateParkStatus
)

from .models import (
    QuanHuyen, PhuongXa, LoaiCongVien, TrangThaiCongVien, CongVien,
    LoaiTienIch, TienIchCongVien, HinhAnhCongVien,
    NhomQuyen, NguoiDung,
    DanhGiaCongVien, LoaiKiemTra, KiemTraCongVien, DanhMucSuCo, BaoCaoSuCo,
    LoaiCay, CayXanh, SuKienCongVien, NhatKyThayDoi, ThongKetruyenCap
)
from .serializers import (
    QuanHuyenSerializer, PhuongXaSerializer, LoaiCongVienSerializer,
    TrangThaiCongVienSerializer, CongVienListSerializer, CongVienDetailSerializer,
    LoaiTienIchSerializer, TienIchCongVienSerializer, HinhAnhCongVienSerializer,
    NhomQuyenSerializer, NguoiDungSerializer, NguoiDungCreateSerializer,
    DanhGiaCongVienSerializer, LoaiKiemTraSerializer, KiemTraCongVienSerializer,
    DanhMucSuCoSerializer, BaoCaoSuCoSerializer,
    LoaiCaySerializer, CayXanhSerializer, SuKienCongVienSerializer,
    NhatKyThayDoiSerializer, ThongKeTruyenCapSerializer
)


class QuanHuyenViewSet(viewsets.ModelViewSet):
    queryset = QuanHuyen.objects.all()
    serializer_class = QuanHuyenSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['ten_quan_huyen', 'ma_code']
    filterset_fields = ['loai']
    
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAdmin()]


class PhuongXaViewSet(viewsets.ModelViewSet):
    queryset = PhuongXa.objects.select_related('ma_quan_huyen')
    serializer_class = PhuongXaSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['ten_phuong_xa', 'ma_code']
    filterset_fields = ['ma_quan_huyen', 'loai']
    
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAdmin()]


class LoaiCongVienViewSet(viewsets.ModelViewSet):
    queryset = LoaiCongVien.objects.all()
    serializer_class = LoaiCongVienSerializer
    
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAdmin()]


class TrangThaiCongVienViewSet(viewsets.ModelViewSet):
    queryset = TrangThaiCongVien.objects.all()
    serializer_class = TrangThaiCongVienSerializer
    
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAdmin()]


class CongVienViewSet(viewsets.ModelViewSet):
    """
    Công viên endpoint.
    
    Quyền:
    - GET: Công khai (AllowAny)
    - POST/DELETE: Chỉ Admin
    - PUT/PATCH: Admin hoặc Manager quản lý công viên đó
    
    LƯU Ý:
    - Manager được edit thông tin công viên (nhưng không được đổi trạng thái/loại)
    - Manager chỉ được edit công viên được gán cho mình
    - Manager quản lý: sự cố, tiện ích, sự kiện của công viên
    """
    queryset = CongVien.objects.select_related(
        'ma_loai', 'ma_trang_thai', 'ma_quan_huyen', 'ma_phuong_xa'
    )
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['ten_cong_vien', 'ma_code', 'dia_chi']
    filterset_fields = ['ma_loai', 'ma_trang_thai', 'ma_quan_huyen', 'da_xac_minh']
    ordering_fields = ['-diem_trung_binh', '-so_luot_danh_gia', 'ten_cong_vien']
    
    def get_permissions(self):
        if self.action == 'tim_gan_nhat':
            return [AllowAny()]
        if self.request.method == 'GET':
            return [AllowAny()]
        # POST/DELETE - Chỉ Admin
        if self.request.method in ['POST', 'DELETE']:
            return [IsAdmin()]
        # PUT/PATCH - Admin hoặc Manager quản lý công viên đó
        if self.request.method in ['PUT', 'PATCH']:
            return [IsManagerOrAdmin()]
        return [IsAdmin()]
    
    def get_serializer_class(self):
        if self.action == 'list':
            return CongVienListSerializer
        return CongVienDetailSerializer

    def _get_request_user(self):
        user = getattr(self.request, 'user', None)
        if getattr(user, 'ma_nguoi_dung', None):
            return user

        token = self.request.headers.get('Authorization', '').replace('Bearer ', '').strip()
        if not token:
            return None

        try:
            user = NguoiDung.objects.select_related('ma_nhom_quyen', 'ma_cong_vien').get(token=token)
            self.request.user = user
            return user
        except NguoiDung.DoesNotExist:
            return None

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.action == 'list':
            queryset = queryset.annotate(tien_ich_so_luong=Count('tien_ich', distinct=True))
        return queryset

    def _tinh_dien_tich_tu_ranh_gioi(self, instance):
        if not HAS_GEOPANDAS or not instance.ranh_gioi:
            return

        try:
            geom = shape(instance.ranh_gioi)
            if not geom.is_valid:
                geom = make_valid(geom)
            
            gdf = gpd.GeoDataFrame({'geometry': [geom]}, crs="EPSG:4326")
            
            target_crs = "EPSG:32648" 
            if hasattr(gdf, 'estimate_utm_crs'):
                try: target_crs = gdf.estimate_utm_crs()
                except: pass 
            
            area_gdf = gdf.to_crs(target_crs)
            area_m2 = area_gdf.geometry.area.iloc[0]
            
            instance.dien_tich_m2 = round(float(area_m2), 2)
            instance.save(update_fields=['dien_tich_m2'])
        except Exception as e:
            print(f"Lỗi tự động tính diện tích: {e}")

    def perform_create(self, serializer):
        instance = serializer.save()
        self._tinh_dien_tich_tu_ranh_gioi(instance)

    def update(self, request, *args, **kwargs):
        """Kiểm tra manager chỉ edit công viên của mình"""
        try:
            park = self.get_object()
            
            # Manager chỉ được edit công viên mình quản lý
            if hasattr(request.user, 'ma_nhom_quyen') and request.user.ma_nhom_quyen.ten_nhom == 'QUAN_LY':
                # FIX: So sánh đúng giữa park.ma_cong_vien (ID) và request.user.ma_cong_vien (FK hoặc ID)
                user_assigned_park_id = request.user.ma_cong_vien_id if hasattr(request.user.ma_cong_vien, 'ma_cong_vien') else request.user.ma_cong_vien
                if park.ma_cong_vien != user_assigned_park_id:
                    return Response(
                        {'error': 'Bạn chỉ được quản lý công viên được gán cho mình'},
                        status=status.HTTP_403_FORBIDDEN
                    )
                
                # Manager KHÔNG được update trạng thái + loại công viên
                update_data = dict(request.data)
                if 'ma_trang_thai' in update_data or 'ma_loai' in update_data:
                    return Response(
                        {'error': 'Bạn không có quyền thay đổi trạng thái hoặc loại công viên'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            
            return super().update(request, *args, **kwargs)
        except Exception as e:
            return Response(
                {'error': f'Lỗi khi cập nhật công viên: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    def perform_update(self, serializer):
        try:
            instance = serializer.save()
            self._tinh_dien_tich_tu_ranh_gioi(instance)
        except Exception as e:
            raise serializers.ValidationError(f'Lỗi khi lưu công viên: {str(e)}')
    
    @action(detail=True, methods=['post'], permission_classes=[IsAdmin], url_path='assign-manager')
    def assign_manager(self, request, pk=None):
        """
        Admin phân công manager cho công viên
        
        Request body: { "manager_id": 123 } hoặc { "manager_username": "manager_test" }
        """
        try:
            park = self.get_object()
            manager_id = request.data.get('manager_id')
            manager_username = request.data.get('manager_username')
            
            # Xác thực đầu vào
            if not manager_id and not manager_username:
                return Response(
                    {'error': 'Cần cung cấp manager_id hoặc manager_username'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Tìm manager
            try:
                if manager_id:
                    manager = NguoiDung.objects.get(ma_nguoi_dung=manager_id)
                else:
                    manager = NguoiDung.objects.get(ten_dang_nhap=manager_username)
            except NguoiDung.DoesNotExist:
                return Response(
                    {'error': 'Manager không tồn tại'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Kiểm tra manager có role QUAN_LY không
            if manager.ma_nhom_quyen.ten_nhom != 'QUAN_LY':
                return Response(
                    {'error': f'Người dùng "{manager.ten_dang_nhap}" không phải là manager'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Kiểm tra manager đã được assign cho công viên khác chưa
            if manager.ma_cong_vien and manager.ma_cong_vien != park:
                return Response(
                    {'error': f'Manager "{manager.ten_dang_nhap}" đã được gán cho công viên khác: {manager.ma_cong_vien.ten_cong_vien}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            existing_manager = NguoiDung.objects.filter(
                ma_cong_vien=park,
                ma_nhom_quyen__ten_nhom='QUAN_LY'
            ).exclude(ma_nguoi_dung=manager.ma_nguoi_dung).first()
            if existing_manager:
                return Response(
                    {'error': f'Công viên "{park.ten_cong_vien}" đã có manager "{existing_manager.ten_dang_nhap}" phụ trách'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Phân công
            manager.ma_cong_vien = park
            manager.save()
            
            return Response({
                'message': f'Đã phân công công viên "{park.ten_cong_vien}" cho manager "{manager.ten_dang_nhap}"',
                'park': CongVienDetailSerializer(park).data,
                'manager': NguoiDungSerializer(manager).data
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {'error': f'Lỗi phân công manager: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'], permission_classes=[IsAdmin], url_path='unassign-manager')
    def unassign_manager(self, request, pk=None):
        """
        Admin gỡ bỏ phân công manager khỏi công viên
        """
        try:
            park = self.get_object()
            
            # Tìm manager được assign cho công viên này
            manager = NguoiDung.objects.filter(ma_cong_vien=park, ma_nhom_quyen__ten_nhom='QUAN_LY').first()
            
            if not manager:
                return Response(
                    {'error': 'Không có manager nào được gán cho công viên này'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Gỡ bỏ phân công
            old_manager_name = manager.ten_dang_nhap
            manager.ma_cong_vien = None
            manager.save()
            
            return Response({
                'message': f'Đã gỡ bỏ manager "{old_manager_name}" khỏi công viên "{park.ten_cong_vien}"',
                'park': CongVienDetailSerializer(park).data
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {'error': f'Lỗi gỡ bỏ manager: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['post', 'get'], url_path='tim-gan-nhat', permission_classes=[AllowAny])
    def tim_gan_nhat(self, request):
        """
        Tìm công viên gần nhất dựa trên vị trí.
        
        Tham số:
        - vi_do/latitude (float): Vĩ độ
        - kinh_do/longitude (float): Kinh độ  
        - ban_kinh_km (float): Bán kính tìm kiếm (km), mặc định 50km
        
        Trả về:
        - count: Số công viên tìm được
        - results: Danh sách TOP 50 công viên gần nhất (chỉ hoạt động)
        - user_location: Vị trí người dùng
        - radius_km: Bán kính tìm kiếm
        """
        vi_do_raw = None
        kinh_do_raw = None
        try:
            data = request.data if request.method == 'POST' else request.query_params
            print(f"🔍 tim_gan_nhat request received - method: {request.method}")
            print(f"📊 Request data/params: {dict(data)}")
            
            def get_val(keys):
                for k in keys:
                    if k in data and data[k] is not None and str(data[k]).strip() != '':
                        return data[k]
                return None

            vi_do_raw = get_val(['vi_do', 'latitude'])
            kinh_do_raw = get_val(['kinh_do', 'longitude'])
            
            if vi_do_raw is None or kinh_do_raw is None:
                print(f"⚠️  Missing coords - lat: {vi_do_raw}, lng: {kinh_do_raw}")
                return Response({'count': 0, 'results': []})

            vi_do = float(vi_do_raw)
            kinh_do = float(kinh_do_raw)
            ban_kinh_km = float(data.get('ban_kinh_km', 50))
            print(f"✅ Coords received - lat: {vi_do}, lng: {kinh_do}, radius: {ban_kinh_km} km")
        except (TypeError, ValueError) as e:
            print(f"❌ Parse error: {e}")
            return Response({
                'error': 'Toa do khong hop le',
                'detail': 'vi_do/latitude va kinh_do/longitude phai la so thuc'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        
        import math
        def haversine(lat1, lon1, lat2, lon2):
            R = 6371
            dLat = math.radians(lat2 - lat1)
            dLon = math.radians(lon2 - lon1)
            a = (math.sin(dLat / 2) * math.sin(dLat / 2) +
                 math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
                 math.sin(dLon / 2) * math.sin(dLon / 2))
            c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
            return R * c

        parks_with_distance = []
        
        # Chỉ tìm công viên "hoạt động" để cho kết quả chính xác
        active_parks = self.queryset.filter(ma_trang_thai__ma_code__iexact='hoat_dong')
        print(f"📍 Total active parks in DB: {active_parks.count()}")

        for park in active_parks:
            if not park.toa_do_trung_tam:
                continue
            
            if isinstance(park.toa_do_trung_tam, list) and len(park.toa_do_trung_tam) == 2:
                try:
                    park_lat, park_lon = float(park.toa_do_trung_tam[0]), float(park.toa_do_trung_tam[1])
                    distance = haversine(vi_do, kinh_do, park_lat, park_lon)
                    print(f"  📌 {park.ten_cong_vien}: {distance:.2f} km")
                    if distance <= ban_kinh_km:
                        park.distance = distance
                        parks_with_distance.append(park)
                except (ValueError, TypeError) as e:
                    print(f"  ⚠️  Error parsing {park.ten_cong_vien}: {e}")
                    continue

        parks_with_distance.sort(key=lambda p: p.distance)
        queryset = parks_with_distance[:50]
        print(f"💚 Parks within {ban_kinh_km}km: {len(queryset)}")
        
        serializer = self.get_serializer(queryset, many=True)
        results = serializer.data
        
        for i, park_data in enumerate(results):
            park_data['khoang_cach_km'] = round(queryset[i].distance, 2)

        return Response({
            'count': len(queryset),
            'user_location': {'latitude': vi_do, 'longitude': kinh_do},
            'radius_km': ban_kinh_km,
            'results': results
        })
    
    @action(detail=False, methods=['get'])
    def ban_do(self, request):
        """
        Endpoint hiển thị bản đồ công viên.
        - Mặc định: hiển thị TẤT CẢ công viên (trừ "quy_hoạch")
        - Query param 'trang_thai' để filter theo trạng thái cụ thể
        - Query param 'include_quy_hoach=true' để bao gồm công viên quy hoạch
        
        Các trạng thái:
        - hoat_dong (xanh): Hoạt động
        - dang_xay_dung (cam): Đang xây dựng
        - cai_tao (vàng): Cải tạo/Sửa chữa
        - tam_dong (đỏ): Tạm đóng
        - ngung_hoat_dong (xám): Ngưng hoạt động
        - quy_hoach (xám): Quy hoạch (không hiển thị mặc định)
        """
        queryset = self.queryset.all()
        
        # Mặc định loại bỏ "quy_hoạch" (chưa xây dựng)
        include_quy_hoach = request.query_params.get('include_quy_hoach', '').lower() == 'true'
        if not include_quy_hoach:
            queryset = queryset.exclude(ma_trang_thai__ma_code__iexact='quy_hoach')
        
        # Filter theo trạng thái nếu có query param
        trang_thai = request.query_params.get('trang_thai')
        if trang_thai:
            queryset = queryset.filter(ma_trang_thai__ma_code__iexact=trang_thai)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def can_kiem_tra(self, request):
        """
        Lấy danh sách công viên cần kiểm tra (chưa kiểm tra trong 30 ngày).
        Chỉ bao gồm công viên có trạng thái "hoạt động".
        """
        thoi_gian_30_ngay = timezone.now() - timedelta(days=30)
        parks_not_checked = CongVien.objects.filter(
            ma_trang_thai__ten_trang_thai='hoat_dong'
        ).exclude(
            kiem_tra__ngay_kiem_tra__gte=thoi_gian_30_ngay
        ).distinct()
        
        serializer = self.get_serializer(parks_not_checked, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def cat_ranh_gioi(self, request, pk=None):
        if not HAS_GEOPANDAS:
            return Response(
                {'error': 'Thư viện Geopandas chưa được cài đặt trên server.'}, 
                status=status.HTTP_501_NOT_IMPLEMENTED
            )

        try:
            park = self.get_object()
            
            if not park.ranh_gioi:
                return Response({'error': 'Công viên chưa có ranh giới để cắt'}, status=400)
            
            if not park.ma_quan_huyen or not park.ma_quan_huyen.hinh_hoc:
                 return Response({'error': 'Quận/Huyện trực thuộc chưa có dữ liệu ranh giới (hinh_hoc)'}, status=400)
            
            park_geom = shape(park.ranh_gioi)
            if not park_geom.is_valid:
                park_geom = make_valid(park_geom)
            gdf_park = gpd.GeoDataFrame({'geometry': [park_geom]}, crs="EPSG:4326")
            
            district_geom = shape(park.ma_quan_huyen.hinh_hoc)
            if not district_geom.is_valid:
                district_geom = make_valid(district_geom)
            gdf_district = gpd.GeoDataFrame({'geometry': [district_geom]}, crs="EPSG:4326")
            
            try:
                clipped_gdf = gpd.clip(gdf_park, gdf_district)
            except Exception as clip_err:
                clipped_geom = park_geom.intersection(district_geom)
                clipped_gdf = gpd.GeoDataFrame({'geometry': [clipped_geom]}, crs="EPSG:4326")
            
            if clipped_gdf.empty or clipped_gdf.geometry.iloc[0].is_empty:
                return Response({
                    'error': 'Ranh giới công viên nằm hoàn toàn bên ngoài ranh giới Quận/Huyện được chọn.'
                }, status=400)
            
            result_geom = clipped_gdf.geometry.iloc[0]
            park.ranh_gioi = mapping(result_geom)
            
            try:
                if not clipped_gdf.crs:
                    clipped_gdf.set_crs("EPSG:4326", allow_override=True, inplace=True)

                target_crs = "EPSG:32648" 
                
                if hasattr(clipped_gdf, 'estimate_utm_crs'):
                    try: target_crs = clipped_gdf.estimate_utm_crs()
                    except: pass # Nếu lỗi estimate thì giữ nguyên 32648
                
                area_gdf = clipped_gdf.to_crs(target_crs)
                area_m2 = area_gdf.geometry.area.iloc[0]
                park.dien_tich_m2 = round(float(area_m2), 2)
            except Exception as calc_err:
                print(f"Lỗi tính diện tích: {calc_err}")
            
            park.save()
            
            return Response({'message': 'Đã cắt ranh giới thành công', 'ranh_gioi': park.ranh_gioi, 'dien_tich_m2': park.dien_tich_m2})
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return Response({'error': str(e)}, status=500)


class LoaiTienIchViewSet(viewsets.ModelViewSet):
    queryset = LoaiTienIch.objects.all()
    serializer_class = LoaiTienIchSerializer
    
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAdmin()]


class TienIchCongVienViewSet(viewsets.ModelViewSet):
    """
    Tiện ích công viên.
    
    Quyền:
    - GET: Công khai (AllowAny)
    - POST/PUT/PATCH: Manager (quản lý công viên) hoặc Admin
    - DELETE: Admin
    """
    queryset = TienIchCongVien.objects.select_related('ma_cong_vien', 'ma_loai_tien_ich')
    serializer_class = TienIchCongVienSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['ma_cong_vien', 'ma_loai_tien_ich', 'tinh_trang']
    
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        if self.request.method in ['POST', 'PUT', 'PATCH', 'DELETE']:
            return [CanManageAmenities()]
        return [IsAdmin()]

    def _get_request_user(self):
        user = getattr(self.request, 'user', None)
        if getattr(user, 'ma_nguoi_dung', None):
            return user

        token = self.request.headers.get('Authorization', '').replace('Bearer ', '').strip()
        if not token:
            return None

        try:
            user = NguoiDung.objects.select_related('ma_nhom_quyen', 'ma_cong_vien').get(token=token)
            self.request.user = user
            return user
        except NguoiDung.DoesNotExist:
            return None

    def get_queryset(self):
        queryset = super().get_queryset()
        request_user = self._get_request_user()
        if request_user and hasattr(request_user, 'ma_nhom_quyen') and request_user.ma_nhom_quyen.ten_nhom == 'QUAN_LY':
            return queryset.filter(ma_cong_vien=request_user.ma_cong_vien)

        return queryset

    def _validate_manager_scope(self, data, instance=None):
        user = getattr(self.request, 'user', None)
        if not user or not hasattr(user, 'ma_nhom_quyen'):
            return None
        if user.ma_nhom_quyen.ten_nhom != 'QUAN_LY':
            return None

        target_park_id = data.get('ma_cong_vien')
        if target_park_id is None and instance is not None:
            target_park_id = instance.ma_cong_vien_id

        if not user.ma_cong_vien_id:
            return Response({'error': 'Manager chưa được gán công viên quản lý.'}, status=status.HTTP_403_FORBIDDEN)

        if str(target_park_id) != str(user.ma_cong_vien_id):
            return Response({'error': 'Bạn chỉ được thao tác tiện ích của công viên được gán.'}, status=status.HTTP_403_FORBIDDEN)
        return None
    
    def create(self, request, *args, **kwargs):
        if hasattr(request.data, 'dict'):
            data = request.data.dict()
        else:
            data = request.data.copy()

        manager_scope_error = self._validate_manager_scope(data)
        if manager_scope_error:
            return manager_scope_error
        
        images = request.FILES.getlist('hinh_anh_files')
        image_urls = []
        
        if images:
            for img in images:
                file_name = default_storage.save(f'amenity_images/{img.name}', ContentFile(img.read()))
                file_url = request.build_absolute_uri(settings.MEDIA_URL + file_name)
                image_urls.append(file_url)
        
        if image_urls:
            data['hinh_anh'] = image_urls
            
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        if hasattr(request.data, 'dict'):
            data = request.data.dict()
        else:
            data = request.data.copy()

        manager_scope_error = self._validate_manager_scope(data, instance=instance)
        if manager_scope_error:
            return manager_scope_error
        
        images = request.FILES.getlist('hinh_anh_files')
        
        if images:
            current_images = instance.hinh_anh or []
            new_urls = []
            for img in images:
                file_name = default_storage.save(f'amenity_images/{img.name}', ContentFile(img.read()))
                file_url = request.build_absolute_uri(settings.MEDIA_URL + file_name)
                new_urls.append(file_url)
            
            data['hinh_anh'] = current_images + new_urls
            
        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        manager_scope_error = self._validate_manager_scope({}, instance=instance)
        if manager_scope_error:
            return manager_scope_error
        return super().destroy(request, *args, **kwargs)


class HinhAnhCongVienViewSet(viewsets.ModelViewSet):
    queryset = HinhAnhCongVien.objects.select_related('ma_cong_vien')
    serializer_class = HinhAnhCongVienSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['ma_cong_vien']
    
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAdmin()]
    
    def create(self, request, *args, **kwargs):
        data = request.data.copy()
        file = request.FILES.get('url_anh')
        
        if file:
            file_name = default_storage.save(f'park_images/{file.name}', ContentFile(file.read()))
            file_url = request.build_absolute_uri(settings.MEDIA_URL + file_name)
            data['url_anh'] = file_url
            
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class NhomQuyenViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = NhomQuyen.objects.all()
    serializer_class = NhomQuyenSerializer


class NguoiDungViewSet(viewsets.ModelViewSet):
    queryset = NguoiDung.objects.select_related('ma_nhom_quyen')
    serializer_class = NguoiDungSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['ten_dang_nhap', 'email', 'ho_ten']
    filterset_fields = ['ma_nhom_quyen', 'dang_hoat_dong']
    
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdmin()]
        return [IsAuthenticated()]
    
    def get_queryset(self):
        if hasattr(self.request, 'user') and self.request.user and self.request.user.ma_nhom_quyen.ten_nhom != 'QUAN_TRI':
            return NguoiDung.objects.filter(ma_nguoi_dung=self.request.user.ma_nguoi_dung)
        return self.queryset
    
    def get_serializer_class(self):
        if self.action == 'create':
            return NguoiDungCreateSerializer
        return NguoiDungSerializer


class DanhGiaCongVienViewSet(viewsets.ModelViewSet):
    queryset = DanhGiaCongVien.objects.select_related('ma_cong_vien', 'ma_nguoi_dung')
    serializer_class = DanhGiaCongVienSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['ma_cong_vien', 'da_duyet']
    
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        if self.request.method in ['PUT', 'PATCH', 'DELETE']:
            return [IsAdmin()]
        return [IsAuthenticated()]
    
    @action(detail=False, methods=['get'], permission_classes=[IsAdmin])
    def danh_gia_chua_duyet(self, request):
        queryset = self.queryset.filter(da_duyet=False)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class LoaiKiemTraViewSet(viewsets.ModelViewSet):
    queryset = LoaiKiemTra.objects.all()
    serializer_class = LoaiKiemTraSerializer
    
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAdmin()]


class KiemTraCongVienViewSet(viewsets.ModelViewSet):
    queryset = KiemTraCongVien.objects.select_related('ma_cong_vien', 'ma_nguoi_kiem_tra')
    serializer_class = KiemTraCongVienSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['ma_cong_vien', 'ngay_kiem_tra__gte', 'ket_qua']
    
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAdmin()]


class DanhMucSuCoViewSet(viewsets.ModelViewSet):
    queryset = DanhMucSuCo.objects.all()
    serializer_class = DanhMucSuCoSerializer
    
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAdmin()]


class BaoCaoSuCoViewSet(viewsets.ModelViewSet):
    """
    Báo cáo sự cố công viên.
    
    Quyền:
    - GET: Công khai (AllowAny)
    - POST: Người dùng xác thực (báo cáo sự cố)
    - PUT/PATCH: Manager (xử lý sự cố của công viên mình) hoặc Admin
    - DELETE: Admin
    """
    queryset = BaoCaoSuCo.objects.select_related(
        'ma_cong_vien', 'ma_danh_muc', 'ma_nguoi_phu_trach', 'ma_nguoi_bao_cao'
    )
    serializer_class = BaoCaoSuCoSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['ma_cong_vien', 'trang_thai', 'muc_do_uu_tien', 'is_archived']
    ordering_fields = ['-muc_do_uu_tien', '-ngay_tao']

    def _get_request_user(self):
        user = getattr(self.request, 'user', None)
        if getattr(user, 'ma_nguoi_dung', None):
            return user

        token = self.request.headers.get('Authorization', '').replace('Bearer ', '').strip()
        if not token:
            return None

        try:
            user = NguoiDung.objects.select_related('ma_nhom_quyen', 'ma_cong_vien').get(token=token)
            self.request.user = user
            return user
        except NguoiDung.DoesNotExist:
            return None
    
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        if self.request.method == 'POST':
            return [CanReportIncident()]  # Người dùng bất kỳ có thể báo cáo
        if self.request.method in ['PUT', 'PATCH']:
            return [CanHandleIncident()]  # Manager/Admin xử lý
        if self.request.method == 'DELETE':
            return [IsAdmin()]
        return [IsAdmin()]
    
    def get_queryset(self):
        """Manager chỉ thấy sự cố của công viên mình"""
        queryset = super().get_queryset()
        is_archived = self.request.query_params.get('is_archived', 'false').lower() == 'true'
        queryset = queryset.filter(is_archived=is_archived)

        request_user = self._get_request_user()
        if request_user and hasattr(request_user, 'ma_nhom_quyen') and request_user.ma_nhom_quyen.ten_nhom == 'QUAN_LY':
            # Manager chỉ thấy sự cố của công viên mình
            return queryset.filter(ma_cong_vien=request_user.ma_cong_vien)
        
        return queryset
    
    def create(self, request, *args, **kwargs):
        if hasattr(request.data, 'dict'):
            data = request.data.dict()
        else:
            data = request.data.copy()
        
        images = request.FILES.getlist('hinh_anh_files')
        image_urls = []
        
        if images:
            for img in images:
                file_name = default_storage.save(f'incident_images/{img.name}', ContentFile(img.read()))
                file_url = request.build_absolute_uri(settings.MEDIA_URL + file_name)
                image_urls.append(file_url)
        
        if image_urls:
            data['url_hinh_anh'] = image_urls
            
        if 'vi_tri' in data and isinstance(data['vi_tri'], str):
            try:
                data['vi_tri'] = json.loads(data['vi_tri'])
            except:
                pass

        if request.user and hasattr(request.user, 'ma_nguoi_dung'):
            data['ma_nguoi_bao_cao'] = request.user.ma_nguoi_dung

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
            
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        """Auto-archive khi cập nhật status thành da_xu_ly"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        # Tạo data dict để có thể modify
        data = dict(request.data)

        if hasattr(request.user, 'ma_nhom_quyen') and request.user.ma_nhom_quyen.ten_nhom == 'QUAN_LY':
            disallowed_fields = set(data.keys()) - {'trang_thai'}
            if disallowed_fields:
                return Response(
                    {'error': 'Manager chi duoc cap nhat trang thai su co cua cong vien duoc giao.'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        # Nếu update status thành "da_xu_ly" → archive
        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        
        if getattr(instance, '_prefetched_objects_cache', None):
            instance._prefetched_objects_cache = {}
        
        return Response(serializer.data)

    @action(detail=True, methods=['patch'])
    def cap_nhat_trang_thai(self, request, pk=None):
        bao_cao = self.get_object()
        trang_thai = request.data.get('trang_thai')
        if trang_thai:
            bao_cao.trang_thai = trang_thai
            
            # Auto-archive khi xử lý xong
            bao_cao.save()
            return Response({'message': 'Cập nhật trạng thái thành công'})
        return Response({'error': 'trang_thai là bắt buộc'}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'], permission_classes=[IsManagerOrAdmin])
    def export_excel(self, request):
        """
        Export incidents to Excel without duplicates.
        Manager: chỉ export sự cố của công viên mình
        Admin: export tất cả
        Query params: ?is_archived=false (default), ?date=YYYY-MM-DD
        """
        try:
            if not HAS_OPENPYXL:
                return Response(
                    {'error': 'openpyxl không được cài đặt'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Get query parameters
            is_archived = request.query_params.get('is_archived', 'false').lower() == 'true'
            date_str = request.query_params.get('date')  # YYYY-MM-DD format for specific day
            
            # Filter incidents
            queryset = self.get_queryset().filter(is_archived=is_archived)
            
            # Manager chỉ export sự cố của công viên mình
            if hasattr(request.user, 'ma_nhom_quyen') and request.user.ma_nhom_quyen.ten_nhom == 'QUAN_LY':
                queryset = queryset.filter(ma_cong_vien=request.user.ma_cong_vien)
            
            # If date specified, filter by the relevant date field for the selected group.
            if date_str:
                try:
                    from datetime import datetime
                    filter_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    date_field = 'ngay_luu_tru__date' if is_archived else 'ngay_tao__date'
                    queryset = queryset.filter(**{date_field: filter_date})
                except:
                    return Response(
                        {'error': 'Invalid date format. Use YYYY-MM-DD'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Remove duplicates: keep only incidents with unique titles per park per day
            seen = set()
            unique_incidents = []
            for incident in queryset:
                key = (incident.ma_cong_vien.ma_cong_vien, incident.tieu_de)
                if key not in seen:
                    seen.add(key)
                    unique_incidents.append(incident)
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Sự cố công viên'
            
            # Define styles
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Header row
            headers = ['STT', 'Công viên', 'Tiêu đề', 'Mô tả', 'Loại sự cố', 'Mức độ ưu tiên',
                       'Trạng thái', 'Người báo cáo', 'Người phụ trách', 'Số xác nhận', 'Ngày tạo']
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            
            # Set column widths
            column_widths = [5, 20, 25, 30, 15, 15, 15, 15, 15, 10, 15]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
            
            # Data rows
            for row_num, incident in enumerate(unique_incidents, 2):
                row_data = [
                    row_num - 1,  # STT
                    incident.ma_cong_vien.ten_cong_vien,
                    incident.tieu_de,
                    incident.noi_dung_mo_ta,
                    incident.ma_danh_muc.ten_danh_muc if incident.ma_danh_muc else 'N/A',
                    incident.get_muc_do_uu_tien_display(),
                    incident.get_trang_thai_display(),
                    incident.ma_nguoi_bao_cao.ten_dang_nhap if incident.ma_nguoi_bao_cao else 'Anonymous',
                    incident.ma_nguoi_phu_trach.ten_dang_nhap if incident.ma_nguoi_phu_trach else 'Chưa phân công',
                    incident.so_nguoi_xac_nhan,
                    incident.ngay_tao.strftime('%d/%m/%Y %H:%M')
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    if col_num == 1:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
            
            # Create response
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            response = HttpResponse(
                excel_file.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="su_co_{date_str or timezone.now().date()}.xlsx"'
            return response
            
        except Exception as e:
            import traceback
            print(f"❌ Export Excel Error: {e}")
            traceback.print_exc()
            return Response(
                {'error': f'Lỗi xuất Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class LoaiCayViewSet(viewsets.ModelViewSet):
    queryset = LoaiCay.objects.all()
    serializer_class = LoaiCaySerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['ten_loai', 'ten_khoa_hoc']

    DEFAULT_TREE_TYPES = [
        {'ten_loai': 'Cay sao den', 'ten_khoa_hoc': 'Hopea odorata', 'mo_ta': 'Loai cay bong mat pho bien tai cong vien va duong pho.'},
        {'ten_loai': 'Cay dau ray', 'ten_khoa_hoc': 'Dipterocarpus alatus', 'mo_ta': 'Cay go lon, tan rong, thuong duoc trong tao canh quan xanh.'},
        {'ten_loai': 'Cay bang lang', 'ten_khoa_hoc': 'Lagerstroemia speciosa', 'mo_ta': 'Loai cay cho hoa tim, phu hop canh quan do thi.'},
        {'ten_loai': 'Cay phuong vi', 'ten_khoa_hoc': 'Delonix regia', 'mo_ta': 'Loai cay hoa do, tao diem nhan mua he cho cong vien.'},
        {'ten_loai': 'Cay lim xet', 'ten_khoa_hoc': 'Peltophorum pterocarpum', 'mo_ta': 'Loai cay bong mat co hoa vang, phat trien nhanh.'},
        {'ten_loai': 'Cay me tay', 'ten_khoa_hoc': 'Samanea saman', 'mo_ta': 'Loai cay co tan rat rong, tao bong mat lon.'},
    ]

    def _ensure_default_tree_types(self):
        if LoaiCay.objects.exists():
            return

        for item in self.DEFAULT_TREE_TYPES:
            LoaiCay.objects.get_or_create(
                ten_loai=item['ten_loai'],
                defaults={
                    'ten_khoa_hoc': item['ten_khoa_hoc'],
                    'mo_ta': item['mo_ta'],
                }
            )
    
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAdmin()]

    def get_queryset(self):
        self._ensure_default_tree_types()
        return super().get_queryset()


class CayXanhViewSet(viewsets.ModelViewSet):
    queryset = CayXanh.objects.select_related('ma_cong_vien', 'ma_loai_cay')
    serializer_class = CayXanhSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['ma_cong_vien', 'ma_loai_cay', 'tinh_trang']
    
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsAdmin()]
    
    @action(detail=False, methods=['get'])
    def thong_ke_tinh_trang(self, request):
        stats = CayXanh.objects.values('tinh_trang').annotate(
            count=Count('ma_cay')
        ).order_by('tinh_trang')
        return Response(stats)


class SuKienCongVienViewSet(viewsets.ModelViewSet):
    """
    Sự kiện công viên.
    
    Quyền:
    - GET: Công khai (AllowAny)
    - POST: Manager (thêm sự kiện cho công viên mình) hoặc Admin
    - PUT/PATCH: Admin (cần để duyệt sự kiện)
    - DELETE: Admin
    """
    queryset = SuKienCongVien.objects.select_related('ma_cong_vien')
    serializer_class = SuKienCongVienSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['ma_cong_vien', 'trang_thai', 'loai_su_kien']
    ordering_fields = ['thoi_gian_bat_dau']

    def _get_request_user(self):
        user = getattr(self.request, 'user', None)
        if getattr(user, 'ma_nguoi_dung', None):
            return user

        token = self.request.headers.get('Authorization', '').replace('Bearer ', '').strip()
        if not token:
            return None

        try:
            user = NguoiDung.objects.select_related('ma_nhom_quyen', 'ma_cong_vien').get(token=token)
            self.request.user = user
            return user
        except NguoiDung.DoesNotExist:
            return None
    
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        if self.request.method in ['POST', 'PUT', 'PATCH', 'DELETE']:
            return [CanCreateEvent()]  # Manager/Admin thao tac su kien
        if self.request.method in ['PUT', 'PATCH']:
            return [IsAdmin()]  # Chi Admin duyet su kien
        if self.request.method == 'DELETE':
            return [IsAdmin()]
        return [IsAdmin()]

    def get_queryset(self):
        queryset = super().get_queryset()

        request_user = self._get_request_user()
        if request_user and hasattr(request_user, 'ma_nhom_quyen') and request_user.ma_nhom_quyen.ten_nhom == 'QUAN_LY':
            return queryset.filter(ma_cong_vien=request_user.ma_cong_vien)

        return queryset

    def create(self, request, *args, **kwargs):
        data = request.data.copy()
        user = getattr(request, 'user', None)

        if user and hasattr(user, 'ma_nhom_quyen') and user.ma_nhom_quyen.ten_nhom == 'QUAN_LY':
            if not user.ma_cong_vien_id:
                return Response({'error': 'Manager chưa được gán công viên quản lý.'}, status=status.HTTP_403_FORBIDDEN)
            if str(data.get('ma_cong_vien')) != str(user.ma_cong_vien_id):
                return Response({'error': 'Manager chỉ được tạo sự kiện cho công viên được gán.'}, status=status.HTTP_403_FORBIDDEN)

        return super().create(request, *args, **kwargs)
    

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        user = self._get_request_user()

        if user and hasattr(user, 'ma_nhom_quyen') and user.ma_nhom_quyen.ten_nhom == 'QUAN_LY':
            if not user.ma_cong_vien_id or instance.ma_cong_vien_id != user.ma_cong_vien_id:
                return Response(
                    {'error': 'Manager chi duoc cap nhat su kien cua cong vien duoc giao.'},
                    status=status.HTTP_403_FORBIDDEN
                )

            if 'da_duyet' in request.data:
                return Response(
                    {'error': 'Manager khong co quyen duyet su kien.'},
                    status=status.HTTP_403_FORBIDDEN
                )

            data = request.data.copy()
            data['ma_cong_vien'] = user.ma_cong_vien_id
            partial = kwargs.pop('partial', False)
            serializer = self.get_serializer(instance, data=data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)

            if getattr(instance, '_prefetched_objects_cache', None):
                instance._prefetched_objects_cache = {}

            return Response(serializer.data)

        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        user = self._get_request_user()

        if user and hasattr(user, 'ma_nhom_quyen') and user.ma_nhom_quyen.ten_nhom == 'QUAN_LY':
            if not user.ma_cong_vien_id or instance.ma_cong_vien_id != user.ma_cong_vien_id:
                return Response(
                    {'error': 'Manager chi duoc xoa su kien cua cong vien duoc giao.'},
                    status=status.HTTP_403_FORBIDDEN
                )

        return super().destroy(request, *args, **kwargs)
    @action(detail=False, methods=['get'])
    def su_kien_sap_toi(self, request):
        """Lấy danh sách sự kiện sắp tới (trong 7 ngày)."""
        thoi_gian_hien_tai = timezone.now()
        thoi_gian_mot_tuan = thoi_gian_hien_tai + timedelta(days=7)
        
        queryset = self.queryset.filter(
            thoi_gian_bat_dau__gte=thoi_gian_hien_tai,
            thoi_gian_bat_dau__lte=thoi_gian_mot_tuan,
            da_duyet=True
        ).order_by('thoi_gian_bat_dau')
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class NhatKyThayDoiViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = NhatKyThayDoi.objects.select_related('ma_nguoi_dung')
    serializer_class = NhatKyThayDoiSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['bang_du_lieu', 'loai_thay_doi']
    ordering_fields = ['-ngay_thay_doi']
    permission_classes = [IsAdmin]


class ThongKeTruyenCapViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ThongKetruyenCap.objects.all()
    serializer_class = ThongKeTruyenCapSerializer
    ordering_fields = ['-ngay']
    permission_classes = [IsAdmin]


@api_view(['GET'])
def dashboard_thong_ke(request):
    try:
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if token:
            try:
                user = NguoiDung.objects.get(token=token)
                if user.ma_nhom_quyen.ten_nhom != 'QUAN_TRI':
                    return Response(
                        {'error': 'Bạn không có quyền truy cập'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            except NguoiDung.DoesNotExist:
                return Response(
                    {'error': 'Token không hợp lệ'},
                    status=status.HTTP_401_UNAUTHORIZED
                )
        
        stats = {
            'tong_cong_vien': CongVien.objects.filter(ma_trang_thai__ten_trang_thai='hoat_dong').count(),
            'cong_vien_can_kiem_tra': CongVien.objects.filter(
                ma_trang_thai__ten_trang_thai='hoat_dong'
            ).exclude(
                kiem_tra__ngay_kiem_tra__gte=timezone.now() - timedelta(days=30)
            ).distinct().count(),
            'baocao_su_co_cho_xu_ly': BaoCaoSuCo.objects.filter(trang_thai='cho_xu_ly').count(),
            'danh_gia_cho_duyet': DanhGiaCongVien.objects.filter(da_duyet=False).count(),
            'tong_dien_tich_m2': CongVien.objects.aggregate(
                total=Sum('dien_tich_m2')
            )['total'] or 0,
            'tong_cay_xanh': CayXanh.objects.count(),
            'cay_benh': CayXanh.objects.filter(tinh_trang='kem').count(),
        }
        return Response(stats)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
